#!/usr/bin/env python3
"""Analyze the three current Excel workbooks that make up the seed database."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / "reports" / "core_excel_database_audit.md"

DOI_RE = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.I)
REFKEY_RE = re.compile(r"\((\d+(?:,\s*\d+)*)\)")


XIAO_COLUMNS = [
    "record_label",
    "notes",
    "citation",
    "diameter_um",
    "diameter_um_err",
    "density_g_cm3",
    "density_g_cm3_err",
    "linear_density_tex",
    "linear_density_tex_err",
    "tenacity_N_tex",
    "tenacity_N_tex_err",
    "tensile_strength_GPa",
    "tensile_strength_GPa_err",
    "initial_modulus_N_tex",
    "initial_modulus_N_tex_err",
    "initial_modulus_GPa",
    "initial_modulus_GPa_err",
    "breaking_strain_pct",
    "breaking_strain_pct_err",
    "rupture_work_J_g",
    "rupture_work_J_g_err",
    "electrical_conductivity_S_cm",
    "electrical_conductivity_S_cm_err",
    "thermal_conductivity_W_mK",
    "thermal_conductivity_W_mK_err",
]

NUMERIC_XIAO_FIELDS = XIAO_COLUMNS[3:]
VALUE_FIELDS = [
    "diameter_um",
    "density_g_cm3",
    "linear_density_tex",
    "tenacity_N_tex",
    "tensile_strength_GPa",
    "initial_modulus_N_tex",
    "initial_modulus_GPa",
    "breaking_strain_pct",
    "rupture_work_J_g",
    "electrical_conductivity_S_cm",
    "thermal_conductivity_W_mK",
]
ERROR_FIELDS = [field for field in NUMERIC_XIAO_FIELDS if field.endswith("_err")]


def clean(value):
    if value is None:
        return pd.NA
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"", "null", "nan", "-", "--"}:
            return pd.NA
        return text
    return value


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def extract_dois(text: str) -> list[str]:
    return sorted({m.group(0).rstrip(" .,)") for m in DOI_RE.finditer(text or "")})


def extract_refkeys(text: str) -> list[str]:
    keys: list[str] = []
    for match in REFKEY_RE.finditer(text or ""):
        keys.extend([part.strip() for part in match.group(1).split(",") if part.strip()])
    return sorted(set(keys))


def classify_xiao(row: pd.Series) -> str:
    label = normalize_text(row["record_label"])
    group = normalize_text(row.get("group_label", ""))
    notes = normalize_text(row["notes"])
    text = f"{label} {group} {notes}"
    if any(token in text for token in ["rice-", "imdea", "pust", "cam-", "utdallas", "mit-", "dwnt", "swnt", "cnt", "this work", "go-", "carbon ", "dr4.7"]):
        return "CNT/CNT hybrid"
    if any(token in text for token in ["toray", "granoc", "dia lead", "dialead", "tenex", "hextow", "hexcel", "t300", "t700", "m35", "m40", "m60", "ysh", "k13", "as4", "im7", "hm63"]):
        return "Carbon fiber comparator"
    if "metal" in text or notes in {"au", "ag", "cu", "ni", "al", "steel"}:
        return "Metal comparator"
    if "glass" in text or "silicon carbide" in text:
        return "Ceramic/glass comparator"
    if any(token in text for token in ["zylon", "dyneema", "hmpe", "polyarylate", "polyaramide", "polymer", "spectra", "vectra", "vectran", "twaron", "technora", "polyester", "pbi", "hs-pe"]):
        return "Polymer/high-performance fiber comparator"
    return "Unclassified"


def read_xiao() -> pd.DataFrame:
    wb = load_workbook(ROOT / "XiaO_DATA.xlsx", data_only=False, read_only=True)
    ws = wb["Sheet1"]
    rows = []
    for excel_row in range(4, ws.max_row + 1):
        values = [clean(ws.cell(excel_row, col).value) for col in range(1, 26)]
        row = dict(zip(XIAO_COLUMNS, values))
        row["source_file"] = "XiaO_DATA.xlsx"
        row["source_sheet"] = "Sheet1"
        row["source_row"] = excel_row
        rows.append(row)
    wb.close()
    df = pd.DataFrame(rows)
    for col in NUMERIC_XIAO_FIELDS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["has_any_measurement"] = df[VALUE_FIELDS].notna().any(axis=1)
    df = df[df["has_any_measurement"]].copy()
    df["group_label"] = df["record_label"].ffill()
    df["material_family"] = df.apply(classify_xiao, axis=1)
    df["citation_text"] = df["citation"].fillna("").astype(str)
    df["doi_list"] = df["citation_text"].map(extract_dois)
    df["doi_count"] = df["doi_list"].map(len)
    df["refkeys"] = (df["record_label"].fillna("").astype(str) + " " + df["notes"].fillna("").astype(str)).map(extract_refkeys)
    return df


def read_radar_sheet(sheet: str) -> pd.DataFrame:
    df = pd.read_excel(ROOT / "RadarFigureSource.xlsx", sheet_name=sheet, header=0, dtype=object)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df.columns = [str(c).strip().replace("\n", " ") for c in df.columns]
    df["source_file"] = "RadarFigureSource.xlsx"
    df["source_sheet"] = sheet
    df["source_row"] = df.index + 2
    text_cols = [c for c in ["Name", "Comments", "Publication"] if c in df.columns]
    df["combined_text"] = df[text_cols].fillna("").astype(str).agg(" ".join, axis=1)
    df["doi_list"] = df["combined_text"].map(extract_dois)
    df["refkeys"] = df["combined_text"].map(extract_refkeys)
    return df


def read_new_g() -> pd.DataFrame:
    df = pd.read_excel(ROOT / "New G fibre table Juan.xlsx", sheet_name="Sheet1", header=2, dtype=object)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df["source_file"] = "New G fibre table Juan.xlsx"
    df["source_sheet"] = "Sheet1"
    df["source_row"] = df.index + 4
    text_cols = [c for c in ["Ref", "Notes"] if c in df.columns]
    df["combined_text"] = df[text_cols].fillna("").astype(str).agg(" ".join, axis=1)
    df["doi_list"] = df["combined_text"].map(extract_dois)
    return df


def count_numeric(df: pd.DataFrame, columns: list[str]) -> list[tuple[str, int, int, float | None, float | None]]:
    out = []
    for col in columns:
        values = pd.to_numeric(df[col], errors="coerce")
        count = int(values.notna().sum())
        out.append(
            (
                col,
                count,
                len(df),
                float(values.min()) if count else None,
                float(values.max()) if count else None,
            )
        )
    return out


def fmt_range(min_v, max_v) -> str:
    if min_v is None:
        return ""
    return f"{min_v:.5g}-{max_v:.5g}"


def md_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(x) for x in row) + " |")
    return "\n".join(lines)


def main() -> None:
    xiao = read_xiao()
    radar = read_radar_sheet("Radar")
    err = read_radar_sheet("Fibers With Error")
    new_g = read_new_g()

    xiao_dois = sorted({doi for dois in xiao["doi_list"] for doi in dois})
    radar_dois = sorted({doi for dois in radar["doi_list"] for doi in dois})
    err_dois = sorted({doi for dois in err["doi_list"] for doi in dois})
    newg_dois = sorted({doi for dois in new_g["doi_list"] for doi in dois})

    xiao_refkeys = sorted({key for keys in xiao["refkeys"] for key in keys})
    radar_refkeys = sorted({key for keys in radar["refkeys"] for key in keys})
    err_refkeys = sorted({key for keys in err["refkeys"] for key in keys})

    overlap_doi_radar = sorted(set(xiao_dois) & set(radar_dois))
    overlap_doi_err = sorted(set(xiao_dois) & set(err_dois))
    overlap_ref_radar = sorted(set(xiao_refkeys) & set(radar_refkeys), key=lambda x: int(x) if x.isdigit() else x)
    overlap_ref_err = sorted(set(xiao_refkeys) & set(err_refkeys), key=lambda x: int(x) if x.isdigit() else x)

    likely_bad_dois = [doi for doi in xiao_dois if doi.startswith("10.1126/science.122806") and doi not in {"10.1126/science.1228061"}]
    tensile_scale_flags = xiao[xiao["tensile_strength_GPa"].gt(100, fill_value=False)].copy()
    tensile_scale_by_family = Counter(tensile_scale_flags["material_family"])

    family_counts = Counter(xiao["material_family"])
    xiao_core_cnt = xiao[xiao["material_family"].eq("CNT/CNT hybrid")].copy()

    property_rows = []
    for field, count, total, min_v, max_v in count_numeric(xiao, VALUE_FIELDS):
        err_field = f"{field}_err"
        err_count = int(pd.to_numeric(xiao[err_field], errors="coerce").notna().sum()) if err_field in xiao else 0
        property_rows.append([field, f"{count}/{total}", f"{err_count}/{total}", fmt_range(min_v, max_v)])

    cnt_property_rows = []
    for field, count, total, min_v, max_v in count_numeric(xiao_core_cnt, VALUE_FIELDS):
        err_field = f"{field}_err"
        err_count = int(pd.to_numeric(xiao_core_cnt[err_field], errors="coerce").notna().sum()) if err_field in xiao_core_cnt else 0
        cnt_property_rows.append([field, f"{count}/{total}", f"{err_count}/{total}", fmt_range(min_v, max_v)])

    duplicate_notes = []
    xiao_keys = set(xiao["record_label"].fillna("").map(normalize_text))
    for sheet_name, df in [("Radar", radar), ("Fibers With Error", err)]:
        for _, row in df.iterrows():
            name = normalize_text(row.get("Name", ""))
            comments = normalize_text(row.get("Comments", ""))
            if name and name in xiao_keys:
                duplicate_notes.append([sheet_name, int(row["source_row"]), row.get("Name", ""), "exact name"])
            elif any(name and name in key for key in xiao_keys):
                duplicate_notes.append([sheet_name, int(row["source_row"]), row.get("Name", ""), "partial name"])
            elif comments and comments in xiao_keys:
                duplicate_notes.append([sheet_name, int(row["source_row"]), row.get("Name", ""), "comment match"])
    duplicate_notes = duplicate_notes[:30]

    lines = []
    lines.append("# Core Excel Database Audit")
    lines.append("")
    lines.append("Generated from `XiaO_DATA.xlsx`, `RadarFigureSource.xlsx`, and `New G fibre table Juan.xlsx`.")
    lines.append("")
    lines.append("## Bottom Line")
    lines.append("")
    lines.append("`XiaO_DATA.xlsx` should be treated as the current primary Excel database. It is a direct Origin/Ashby-style table with paired value/error columns, citation text, geometry, density, linear density, mechanical properties, electrical conductivity, and thermal conductivity. The older radar workbook is useful for derived radar plotting and uncertainty cross-checks, but it is incomplete as a database source.")
    lines.append("")
    lines.append("## Xiao Workbook Structure")
    lines.append("")
    lines.append(f"- Sheet: `Sheet1`")
    lines.append(f"- Raw worksheet size: `A1:Y124`")
    lines.append(f"- Data rows with at least one measurement: {len(xiao)}")
    lines.append(f"- Unique DOI-like strings: {len(xiao_dois)}")
    lines.append(f"- Rows with at least one DOI-like string: {int(xiao['doi_count'].gt(0).sum())}/{len(xiao)}")
    lines.append(f"- Rows without DOI-like strings: {int(xiao['doi_count'].eq(0).sum())}/{len(xiao)}")
    lines.append("")
    lines.append("### Xiao Material-Family Counts")
    lines.append("")
    lines.append(md_table(["Material family", "Rows"], [[k, v] for k, v in family_counts.most_common()]))
    lines.append("")
    lines.append("### Xiao Property Coverage: All Rows")
    lines.append("")
    lines.append(md_table(["Field", "Value coverage", "Error coverage", "Range"], property_rows))
    lines.append("")
    lines.append("### Xiao Property Coverage: CNT/CNT-Hybrid Rows")
    lines.append("")
    lines.append(md_table(["Field", "Value coverage", "Error coverage", "Range"], cnt_property_rows))
    lines.append("")
    lines.append("## Citation/Reference Coverage")
    lines.append("")
    lines.append(md_table(
        ["Source", "Rows", "Unique DOI-like strings", "Reference keys"],
        [
            ["XiaO_DATA.xlsx", len(xiao), len(xiao_dois), len(xiao_refkeys)],
            ["RadarFigureSource.xlsx / Radar", len(radar), len(radar_dois), len(radar_refkeys)],
            ["RadarFigureSource.xlsx / Fibers With Error", len(err), len(err_dois), len(err_refkeys)],
            ["New G fibre table Juan.xlsx", len(new_g), len(newg_dois), "n/a"],
        ],
    ))
    lines.append("")
    lines.append("### DOI Overlap")
    lines.append("")
    lines.append(f"- Xiao ∩ Radar DOI-like strings: {len(overlap_doi_radar)}")
    lines.append(f"- Xiao ∩ Fibers With Error DOI-like strings: {len(overlap_doi_err)}")
    lines.append(f"- Xiao ∩ New G DOI-like strings: {len(set(xiao_dois) & set(newg_dois))}")
    lines.append("")
    lines.append("### Reference-Key Overlap")
    lines.append("")
    lines.append(f"- Xiao ∩ Radar reference keys: {', '.join(overlap_ref_radar) if overlap_ref_radar else 'none'}")
    lines.append(f"- Xiao ∩ Fibers With Error reference keys: {', '.join(overlap_ref_err) if overlap_ref_err else 'none'}")
    lines.append("")
    lines.append("### DOI Strings Requiring Validation")
    lines.append("")
    if likely_bad_dois:
        lines.append("These DOI-like strings appear in comparator rows and look suspicious because they are sequential variants of one Science DOI. Treat them as unverified until Crossref validation:")
        lines.append("")
        for doi in likely_bad_dois:
            lines.append(f"- `{doi}`")
    else:
        lines.append("No obvious suspicious DOI-like strings detected by the simple heuristic.")
    lines.append("")
    lines.append("## Unit/Scale Flags")
    lines.append("")
    lines.append(f"- Rows with `tensile_strength_GPa > 100`: {len(tensile_scale_flags)}")
    if len(tensile_scale_flags):
        lines.append("- These are concentrated in comparator rows and are almost certainly MPa-scale values despite the exported column label `GPa`.")
        lines.append("- Example: Toray T1100GC has density 1.79 g cm-3, tenacity 3.91 N tex-1, and tensile strength 7000. This is consistent with 7000 MPa = 7 GPa, because 7/1.79 = 3.91.")
        lines.append("- Import rule: preserve the raw field, then create a canonical `tensile_strength_GPa_canonical`; for comparator rows with raw values >100, divide by 1000 and flag `unit_inferred_from_tenacity_density`.")
        lines.append("")
        lines.append(md_table(["Material family", "Flagged rows"], [[k, v] for k, v in tensile_scale_by_family.most_common()]))
    lines.append("")
    lines.append("## Overlap With Older Excel Workbooks")
    lines.append("")
    lines.append("The older radar workbook appears to be a derived/partial plotting workbook, not the primary source. It overlaps with Xiao by DOI/reference key for many CNT-fiber rows, but it also contains formula-derived radar fields, summary/max rows, and incomplete citation cells.")
    lines.append("")
    if duplicate_notes:
        lines.append("Examples of likely row-level overlap by name/comment:")
        lines.append("")
        lines.append(md_table(["Workbook sheet", "Row", "Name", "Match type"], duplicate_notes))
        lines.append("")
    lines.append("`New G fibre table Juan.xlsx` is a small 3-row addendum for graphene fibers. It overlaps conceptually with the lower addendum embedded in `FibersStrengthYr`, but not with most Xiao rows.")
    lines.append("")
    lines.append("## Mapping To Website Data Model")
    lines.append("")
    lines.append("### publication")
    lines.append("")
    lines.append("- Use Xiao `Citation` as the primary raw citation field.")
    lines.append("- Extract DOI-like strings, URLs, free-text citation fragments, and reference keys separately.")
    lines.append("- Do not mark a DOI valid until Crossref/OpenAlex validation.")
    lines.append("")
    lines.append("### sample")
    lines.append("")
    lines.append("- `record_label` and `notes` contain sample/material identity.")
    lines.append("- Current classification can start with: CNT/CNT hybrid, carbon fiber comparator, polymer/high-performance fiber comparator, ceramic/glass comparator, metal comparator.")
    lines.append("- CNT-specific fields such as CNT type, synthesis method, heat treatment, densification/stretching, and hybrid/additive should be parsed from `record_label`/`notes` but manually reviewed.")
    lines.append("")
    lines.append("### measurement")
    lines.append("")
    lines.append("- Xiao directly maps to diameter, density, linear density, tenacity, tensile strength, initial modulus, breaking strain, rupture work, electrical conductivity, and thermal conductivity.")
    lines.append("- Specific volume and specific electrical conductivity can be calculated from density/electrical conductivity while preserving original values.")
    lines.append("- G:D ratio is not present in Xiao and must remain sourced from the radar workbook or future literature extraction.")
    lines.append("")
    lines.append("### conditions")
    lines.append("")
    lines.append("- Xiao contains limited condition metadata.")
    lines.append("- Heat-treatment temperature can be parsed from labels like `S&DWNT-2700℃`.")
    lines.append("- Measurement method, atmosphere, gauge length, and strain rate are mostly absent and should be nullable fields plus quality/completeness badges.")
    lines.append("")
    lines.append("### provenance")
    lines.append("")
    lines.append("- Preserve `source_file`, `source_sheet`, and `source_row` for every imported row.")
    lines.append("- Add `source_export = OriginPro/Xiao Ashby export` for Xiao rows.")
    lines.append("- Add `extraction_method = exported_table` for Xiao, `manual_addendum` for New G, and `derived_plot_workbook` for Radar unless verified otherwise.")
    lines.append("")
    lines.append("### status")
    lines.append("")
    lines.append("- Public records should still expose only official/verified rows.")
    lines.append("- Internal statuses remain necessary for DOI validation, unit review, duplicate resolution, and source-condition review.")
    lines.append("")
    lines.append("## Suggested Import Order")
    lines.append("")
    lines.append("1. Import Xiao rows as the primary table with row-level provenance.")
    lines.append("2. Normalize units and create canonical measurement fields.")
    lines.append("3. Validate DOI-like strings through Crossref/OpenAlex and split citation text into DOI, URL, journal/title/year where possible.")
    lines.append("4. Deduplicate against `RadarFigureSource.xlsx`; keep Radar-derived values only where Xiao lacks a field, especially G:D ratio and existing radar-specific derived columns.")
    lines.append("5. Add `New G fibre table Juan.xlsx` as a separate addendum table after checking whether the 2025 unpublished row should be internal-only.")
    lines.append("6. Only then build the website data API and comparison-level logic.")
    lines.append("")

    REPORT.parent.mkdir(exist_ok=True)
    REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(REPORT)


if __name__ == "__main__":
    main()
