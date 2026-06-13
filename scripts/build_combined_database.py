#!/usr/bin/env python3
"""Build a combined seed database from the three Excel workbooks.

Outputs:
  data/processed/combined_records.csv
  data/processed/measurements_long.csv
  data/processed/publications.csv
  data/processed/source_issues.csv
  data/processed/data_dictionary.csv
  data/processed/cnt_property_database.sqlite
"""

from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "processed"

DOI_RE = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.I)
URL_RE = re.compile(r"https?://\S+|www\.\S+", re.I)
REFKEY_RE = re.compile(r"\((\d+(?:,\s*\d+)*)\)")

NATURE_2025 = {
    "doi": "10.1038/s41563-025-02384-7",
    "url": "https://www.nature.com/articles/s41563-025-02384-7",
    "title": "High-performance graphene-based carbon fibres prepared at room temperature via domain folding",
    "journal": "Nature Materials",
    "year": 2026,
    "published_date": "2025-10-20",
    "issue": "25, 191-198",
    "authors_short": "Peng Li et al.",
}

JAMES_META_2021 = {
    "source_file": "20210409_Metaanalysis_database.xlsx",
    "doi": "10.1002/adma.202008432",
    "title": "A Meta-Analysis of Conductive and Strong Carbon Nanotube Materials",
    "journal": "Advanced Materials",
    "year": 2021,
    "authors_short": "John S. Bulmer et al.",
    "source_export": "Bulmer/James CNT meta-analysis supporting workbook",
}

EXCLUDED_ROWS: list[dict[str, Any]] = []

DOI_OVERRIDES = {
    "https://www.sciencedirect.com/science/article/pii/S0008622316310752": "10.1016/j.carbon.2016.12.006",
    "https://www.sciencedirect.com/science/article/pii/S0008622321004140": "10.1016/j.carbon.2021.04.033",
    "https://www.sciencedirect.com/science/article/pii/S0008622322003189": "10.1016/j.carbon.2022.04.040",
    "cho, y. s. et al. superstrong carbon nanotube yarns": "10.1002/advs.202204250",
}

BEHABTU_SCIENCE_DOI = "10.1126/science.1228061"
BEHABTU_FAKE_METAL_DOIS = {
    "10.1126/science.1228062": BEHABTU_SCIENCE_DOI,
    "10.1126/science.1228063": BEHABTU_SCIENCE_DOI,
    "10.1126/science.1228064": BEHABTU_SCIENCE_DOI,
    "10.1126/science.1228065": BEHABTU_SCIENCE_DOI,
}

COMMERCIAL_SOURCE_TOKENS = [
    "matweb.com",
    "cf-composites.toray",
    "ngfworld.com",
    "mccfc.com",
    "teijincarbon.com",
    "hexcel.com",
    "honeywell.com/spectra",
    "toyobo.cn",
    "epsilon-composite.com",
]

CANONICAL_DUPLICATE_FIELDS = [
    "density_kg_m3",
    "diameter_m",
    "specific_strength_N_m_kg",
    "tensile_strength_Pa",
    "specific_modulus_N_m_kg",
    "initial_modulus_Pa",
    "breaking_strain_fraction",
    "rupture_work_J_kg",
    "electrical_conductivity_S_m",
    "specific_electrical_conductivity_S_m2_kg",
    "thermal_conductivity_W_mK",
    "specific_thermal_conductivity_W_m2_K_kg",
    "g_d_ratio",
    "ampacity_A_m2",
]


XIAO_COLUMNS = [
    "record_label",
    "notes",
    "citation_raw",
    "diameter_um_raw",
    "diameter_um_error_raw",
    "density_g_cm3_raw",
    "density_g_cm3_error_raw",
    "linear_density_tex_raw",
    "linear_density_tex_error_raw",
    "tenacity_N_tex_raw",
    "tenacity_N_tex_error_raw",
    "tensile_strength_GPa_raw",
    "tensile_strength_GPa_error_raw",
    "initial_modulus_N_tex_raw",
    "initial_modulus_N_tex_error_raw",
    "initial_modulus_GPa_raw",
    "initial_modulus_GPa_error_raw",
    "breaking_strain_pct_raw",
    "breaking_strain_pct_error_raw",
    "rupture_work_J_g_raw",
    "rupture_work_J_g_error_raw",
    "electrical_conductivity_S_cm_raw",
    "electrical_conductivity_S_cm_error_raw",
    "thermal_conductivity_W_mK_raw",
    "thermal_conductivity_W_mK_error_raw",
]


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"", "null", "nan", "-", "--"}:
            return None
        return text
    return value


def clean_reported(value: Any) -> Any:
    value = clean(value)
    if isinstance(value, str) and value.lower() in {
        "not reported",
        "not reported/ can not calculate",
        "not reported/can not calculate",
        "can not calculate",
        "cannot calculate",
        "nan",
    }:
        return None
    return value


def dedupe_headers(headers: list[Any]) -> list[str]:
    """Preserve duplicate Excel headers as pandas-style names."""
    counts: dict[str, int] = {}
    out: list[str] = []
    for idx, header in enumerate(headers, start=1):
        base = clean(header) or f"Unnamed: {idx}"
        base = str(base)
        seen = counts.get(base, 0)
        out.append(base if seen == 0 else f"{base}.{seen}")
        counts[base] = seen + 1
    return out


def to_float(value: Any) -> float | None:
    value = clean_reported(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def norm_text(*values: Any) -> str:
    parts = [str(v) for v in values if clean(v) is not None]
    return re.sub(r"\s+", " ", " ".join(parts).strip())


def lower_text(*values: Any) -> str:
    return norm_text(*values).lower()


def extract_dois(text: str | None) -> list[str]:
    return sorted({m.group(0).rstrip(" .,)") for m in DOI_RE.finditer(text or "")})


def extract_urls(text: str | None) -> list[str]:
    return sorted({m.group(0).rstrip(" .,)") for m in URL_RE.finditer(text or "")})


def extract_refkeys(text: str | None) -> list[str]:
    out: list[str] = []
    for match in REFKEY_RE.finditer(text or ""):
        out.extend(part.strip() for part in match.group(1).split(",") if part.strip())
    return sorted(set(out), key=lambda x: int(x) if x.isdigit() else x)


def doi_overrides_for(text: str | None) -> list[str]:
    source = (text or "").lower()
    dois = []
    for needle, doi in DOI_OVERRIDES.items():
        if needle.lower() in source:
            dois.append(doi)
    return sorted(set(dois))


def correct_known_citation_errors(text: str | None, material_family: str) -> str:
    source = norm_text(text)
    if material_family == "metal_comparator" and "matweb.com" in source.lower():
        for bad_doi, corrected_doi in BEHABTU_FAKE_METAL_DOIS.items():
            source = re.sub(re.escape(bad_doi), corrected_doi, source, flags=re.I)
    return source


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return f"{prefix}_{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:12]}"


def classify_material(label: Any, notes: Any, group: Any = None, source_file: str = "", citation: Any = None) -> str:
    text = lower_text(label, notes, group, citation, source_file)
    label_text = lower_text(label)
    if is_derived_label(label, notes):
        return "derived_plot_scaffold"
    if any(token in text for token in ["cnt/cu", "cnts/cu", "mwcnt/cu", "carbon nanotube/copper", "carbon nanotube copper", "cnt copper", "cu composite wire"]):
        return "CNT_metal_composite"
    if "graphene" in text or re.search(r"\bgo[-: ]", text):
        return "graphene_or_GO_fiber"
    if any(token in text for token in ["rice-", "imdea", "pust", "cam-", "utdallas", "mit-", "dwnt", "swnt", "mwnt", "cnt", "this work", "dr4.7", "carbon nanotube", "nanotube yarn", "s·dw", "s&dw", "s&dwnt", "sw:", "dw:", "pasqualli"]) or label_text in {"carbon", "vilatela"}:
        return "CNT_or_CNT_hybrid"
    if any(token in text for token in ["toray", "granoc", "dialead", "tenex", "hextow", "hexcel", "t300", "t700", "m35", "m40", "m60", "ysh", "k13", "as4", "im7", "hm63", "epsilon", "coal tar pitch"]):
        return "carbon_fiber_comparator"
    if any(token in text for token in ["metal", "copper", "aluminum", "aluminium", "steel", "aermet", "astm a36"]) or any(token == text.strip() for token in ["au", "ag", "cu", "ni", "al"]):
        return "metal_comparator"
    if "glass" in text or "silicon carbide" in text:
        return "ceramic_or_glass_comparator"
    if any(token in text for token in ["zylon", "dyneema", "hmpe", "polyarylate", "polyaramide", "polymer", "spectra", "vectra", "vectran", "twaron", "technora", "polyester", "pbi", "hs-pe"]):
        return "polymer_fiber_comparator"
    return "unclassified"


def is_derived_label(label: Any, notes: Any) -> bool:
    text = lower_text(label, notes)
    if not text:
        return True
    return bool(
        re.search(r"\b(max|normalized by|score|comments|label|name|data)\b", text)
        or text.startswith("density ")
        or text.startswith("tenacity ")
    )


def excluded_row(source_file: str, source_sheet: str, source_row: int, reason: str, raw: dict[str, Any]) -> dict[str, Any]:
    return {
        "excluded_id": stable_id("excluded", source_file, source_sheet, source_row, reason),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "reason": reason,
        "raw_payload_json": json.dumps(raw, ensure_ascii=False, default=str),
    }


def radar_exclusion_reason(raw: dict[str, Any]) -> str | None:
    label = raw.get("Name")
    notes = raw.get("Comments")
    if clean(label) is None and clean(notes) is None:
        return "unlabeled_plot_cell"
    if is_derived_label(label, notes):
        return "derived_plot_header_score_or_maximum"
    return None


def infer_form_factor(material_family: str, label: Any, notes: Any) -> str:
    text = lower_text(label, notes)
    family = material_family.lower()
    if "individual" in text:
        return "individual_nanotube_or_bundle"
    if re.search(r"\b(sheets?|films?|mats?)\b", text):
        return "sheet_mat_film"
    if re.search(r"\b(foams?|aerogels?)\b", text):
        return "foam_aerogel"
    if re.search(r"\b(forests?|arrays?)\b", text):
        return "forest_array"
    if re.search(r"\bbucky(papers?)?\b", text):
        return "buckypaper"
    if family.endswith("comparator") or "fiber" in family or "cnt" in family or material_family == "CNT_metal_composite":
        return "fiber_yarn"
    return "unknown"


def infer_cnt_type(label: Any, notes: Any) -> str | None:
    text = lower_text(label, notes)
    if "s&dwnt" in text or "s·dw" in text or "s&dw" in text:
        return "SWCNT_DWCNT_mix"
    if "dwnt" in text or re.search(r"\bdw[:\-]?", text):
        return "DWCNT"
    if "swnt" in text or re.search(r"\bsw[:\-]?", text):
        return "SWCNT"
    if "mwnt" in text:
        return "MWCNT"
    return None


def infer_synthesis(label: Any, notes: Any, material_family: str) -> str | None:
    text = lower_text(label, notes)
    if material_family != "CNT_or_CNT_hybrid":
        return None
    if any(token in text for token in ["rice", "solution-spun", "solution spun", "ss-", "h-ss"]):
        return "solution_spun"
    if any(token in text for token in ["imdea", "pust", "cam", "direct", "dr4.7"]):
        return "direct_spun_FCCVD"
    if any(token in text for token in ["utdallas", "mit", "va-"]):
        return "forest_drawn_or_VA_CNT"
    return None


def infer_postprocessing(label: Any, notes: Any) -> str | None:
    text = lower_text(label, notes)
    tags = []
    if "raw" in text:
        tags.append("raw")
    if "den" in text or "dens" in text:
        tags.append("densified")
    if "fully dd" in text or "double-drawn" in text:
        tags.append("fully_double_drawn")
    if "heat" in text or re.search(r"\d{4}\s*℃", text):
        tags.append("heat_treated")
    if "go-" in text or "go:" in text:
        tags.append("GO_hybrid")
    if "csa" in text:
        tags.append("CSA_processed")
    return ";".join(tags) if tags else None


def infer_heat_temp(label: Any, notes: Any) -> float | None:
    match = re.search(r"(\d{3,4})\s*℃", norm_text(label, notes))
    return float(match.group(1)) if match else None


def comparison_scope(material_family: str, form_factor: str) -> str:
    if material_family == "CNT_or_CNT_hybrid" and form_factor == "fiber_yarn":
        return "strict_candidate_missing_conditions"
    if material_family in {"CNT_or_CNT_hybrid", "graphene_or_GO_fiber", "CNT_metal_composite"}:
        return "normalized_candidate"
    return "context_comparator"


def james_material_family(category: Any, category2: Any, reference: Any, notes: Any) -> str:
    category_text = lower_text(category, category2, notes)
    reference_text = lower_text(reference)
    text = lower_text(category, category2, reference, notes)
    category2_text = lower_text(category2)
    if "metal cnt composite" in category_text or "cnt/cu" in text or "nanotube/copper" in text:
        return "CNT_metal_composite"
    if category2_text == "metal" or re.search(r"\bmetal\b", category_text) or any(token in category_text for token in ["aluminum", "aluminium", "copper", "steel", "silver", "gold", "nickel"]):
        return "metal_comparator"
    if "graphene" in category_text:
        return "graphene_or_GO_fiber"
    if "carbon fiber" in category_text:
        return "carbon_fiber_comparator"
    if "polymer" in category_text or "synthetic fiber" in category_text or "paper" in category_text:
        return "polymer_fiber_comparator"
    if any(token in category_text for token in ["gic", "graphite", "glassy carbon", "amorphous carbon", "diamond"]):
        return "other_carbon_comparator"
    if category2_text == "cnt" or "cnt" in category_text or "nanotube" in category_text:
        return "CNT_or_CNT_hybrid"
    if "graphene" in reference_text:
        return "graphene_or_GO_fiber"
    if any(token in reference_text for token in ["carbon nanotube", "nanotube fiber", "cnt fiber", "cnt yarn"]):
        return "CNT_or_CNT_hybrid"
    return classify_material(reference, notes, category, JAMES_META_2021["source_file"], reference)


def james_form_factor(category: Any, reference: Any, notes: Any, material_family: str) -> str:
    text = lower_text(category, reference, notes)
    if "individual" in text:
        return "individual_nanotube_or_bundle"
    if "film" in text or "sheet" in text or "mat" in text or "paper" in text:
        if "bucky" in text:
            return "buckypaper"
        return "sheet_mat_film"
    if "bulk material" in text:
        return "bulk"
    return infer_form_factor(material_family, reference, notes)


def james_cnt_type(category: Any, reference: Any, notes: Any) -> str | None:
    text = lower_text(category, reference, notes)
    if "few-wall" in text or "few wall" in text or "fwcnt" in text:
        return "FWCNT"
    if "double-wall" in text or "double wall" in text or "dwcnt" in text:
        return "DWCNT"
    if "single-wall" in text or "single wall" in text or "swcnt" in text or "swnt" in text:
        return "SWCNT"
    if "multiwall" in text or "multi-wall" in text or "mwcnt" in text or "mwnt" in text:
        return "MWCNT"
    return None


def james_synthesis(row: dict[str, Any], material_family: str) -> str | None:
    if material_family not in {"CNT_or_CNT_hybrid", "CNT_metal_composite"}:
        return None
    process = lower_text(row.get("Production Process"), row.get("Reference"), row.get("Notes"))
    if "wet" in process:
        return "wet_spun"
    if "floating" in process or "direct" in process or "aerosol" in process:
        return "direct_spun_FCCVD"
    if "array" in process or "forest" in process or "draw" in process:
        return "forest_drawn_or_VA_CNT"
    if "acid" in process or "superacid" in process:
        return "solution_spun"
    return clean(row.get("Production Process"))


def james_postprocessing(row: dict[str, Any]) -> str | None:
    tags = []
    for field in ["Post Process", "Intentionally added intercalation dope", "Doped or Acid Exposure (Yes/ No)", "Notes", "Reference"]:
        text = lower_text(row.get(field))
        if not text:
            continue
        if "iodine" in text or "icl" in text or "bromine" in text or "doped" in text or "dope" in text:
            tags.append("doped_or_intercalated")
        if "acid" in text:
            tags.append("acid_exposed")
        if "stretch" in text or "draw" in text:
            tags.append("drawn_or_stretched")
        if "dens" in text:
            tags.append("densified")
        if "anneal" in text or "heat" in text:
            tags.append("heat_treated")
    tags = list(dict.fromkeys(tags))
    explicit = clean(row.get("Post Process"))
    if explicit and explicit not in tags:
        tags.append(str(explicit))
    return ";".join(tags) if tags else None


def has_commercial_source(text: str | None) -> bool:
    source = (text or "").lower()
    return any(token in source for token in COMMERCIAL_SOURCE_TOKENS)


def evidence_classification(material_family: str, citation_text: str | None, dois: list[str], urls: list[str]) -> tuple[str, str, str]:
    has_doi = bool(dois)
    has_url = bool(urls)
    commercial = has_commercial_source(citation_text)
    comparator = material_family.endswith("_comparator")

    if comparator and commercial and has_doi:
        return (
            "mixed_peer_reviewed_and_commercial_comparator",
            "context_benchmark_not_primary_literature",
            "Comparator benchmark compiled from a peer-reviewed Ashby/comparison figure plus a commercial/material database source; show with a benchmark-source badge.",
        )
    if comparator and commercial:
        return (
            "commercial_or_web_specsheet_comparator",
            "context_benchmark_non_peer_reviewed",
            "Comparator benchmark from manufacturer, MatWeb, or other web/spec-sheet source; cite as non-peer-reviewed contextual benchmark.",
        )
    if comparator and has_doi:
        return (
            "peer_reviewed_comparator",
            "context_benchmark_peer_reviewed",
            "Comparator benchmark from peer-reviewed literature; not a CNT measurement.",
        )
    if material_family in {"CNT_or_CNT_hybrid", "graphene_or_GO_fiber", "CNT_metal_composite"} and has_doi:
        return (
            "peer_reviewed_research_record",
            "primary_literature_candidate",
            "Research-material record with DOI; still requires sample, unit, and condition curation before public use.",
        )
    if material_family in {"CNT_or_CNT_hybrid", "graphene_or_GO_fiber", "CNT_metal_composite"} and has_url:
        return (
            "url_only_research_record",
            "needs_primary_literature_review",
            "Research-material record has only a URL/source text; find DOI or primary publication before public use.",
        )
    if has_doi:
        return (
            "peer_reviewed_source",
            "literature_candidate",
            "Record has DOI-backed provenance but still requires curation.",
        )
    return (
        "unresolved_or_internal_source",
        "needs_source_review",
        "Record lacks DOI-backed provenance; keep internal until source is resolved.",
    )


def base_record(source_file: str, source_sheet: str, source_row: int, label: Any, notes: Any, citation: Any, group: Any = None) -> dict[str, Any]:
    material_family = classify_material(label, notes, group, source_file, citation)
    form_factor = infer_form_factor(material_family, label, notes)
    citation_text = correct_known_citation_errors(norm_text(citation), material_family)
    dois = sorted(set(extract_dois(citation_text) + doi_overrides_for(citation_text)))
    urls = extract_urls(citation_text)
    refkeys = extract_refkeys(norm_text(label, notes))
    source_citation_class, evidence_tier, comparison_note = evidence_classification(material_family, citation_text, dois, urls)
    record = {
        "record_id": stable_id("rec", source_file, source_sheet, source_row, label, notes),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "source_priority": {
            "literature_addendum_records.tsv": 0,
            "XiaO_DATA.xlsx": 1,
            "RadarFigureSource.xlsx": 2,
            "New G fibre table Juan.xlsx": 3,
            JAMES_META_2021["source_file"]: 4,
        }.get(source_file, 9),
        "record_label": clean(label),
        "group_label": clean(group),
        "sample_name": clean(notes) or clean(label),
        "citation_raw": citation_text or None,
        "doi_raw": ";".join(dois) if dois else None,
        "url_raw": ";".join(urls) if urls else None,
        "reference_keys": ";".join(refkeys) if refkeys else None,
        "source_citation_raw": None,
        "secondary_source_doi_raw": None,
        "secondary_source_title": None,
        "secondary_source_authors_short": None,
        "secondary_source_journal": None,
        "secondary_source_year": None,
        "original_reference_raw": None,
        "doi_resolution_status": None,
        "doi_resolution_score": None,
        "duplicate_of_record_id": None,
        "duplicate_match_score": None,
        "publication_title": None,
        "publication_authors_short": None,
        "publication_journal": None,
        "publication_year": None,
        "publication_published_date": None,
        "publication_issue_pages": None,
        "material_family": material_family,
        "form_factor": form_factor,
        "cnt_type": infer_cnt_type(label, notes),
        "synthesis_method": infer_synthesis(label, notes, material_family),
        "postprocessing": infer_postprocessing(label, notes),
        "heat_treatment_C": infer_heat_temp(label, notes),
        "condition_temperature_C": None,
        "condition_atmosphere": None,
        "measurement_method": None,
        "gauge_length_mm": None,
        "strain_rate_s_inv": None,
        "provenance_table_figure_page": None,
        "source_export": {
            "literature_addendum_records.tsv": "manual literature addendum from local PDFs",
            "XiaO_DATA.xlsx": "OriginPro export from Xiao Ashby plot workbook",
            "RadarFigureSource.xlsx": "derived radar/plot workbook",
            "New G fibre table Juan.xlsx": "manual graphene-fiber addendum",
            JAMES_META_2021["source_file"]: JAMES_META_2021["source_export"],
        }.get(source_file, None),
        "extraction_method": {
            "literature_addendum_records.tsv": "manual_literature_extract",
            "XiaO_DATA.xlsx": "exported_table",
            "RadarFigureSource.xlsx": "derived_plot_workbook",
            "New G fibre table Juan.xlsx": "manual_addendum",
            JAMES_META_2021["source_file"]: "peer_reviewed_meta_analysis_workbook",
        }.get(source_file, "unknown"),
        "comparison_scope": comparison_scope(material_family, form_factor),
        "source_citation_class": source_citation_class,
        "evidence_tier": evidence_tier,
        "comparison_note": comparison_note,
        "public_status": "internal_seed_not_public",
        "internal_status": "needs_reference_unit_condition_review",
        "curator": "source_compilation",
        "confidence_score": None,
        "raw_payload_json": None,
    }
    if source_file == "New G fibre table Juan.xlsx" and "to be published" in lower_text(citation):
        record.update(
            {
                "citation_raw": f"{NATURE_2025['title']}. {NATURE_2025['journal']} ({NATURE_2025['published_date']}). {NATURE_2025['url']}",
                "doi_raw": NATURE_2025["doi"],
                "url_raw": NATURE_2025["url"],
                "publication_title": NATURE_2025["title"],
                "publication_authors_short": NATURE_2025["authors_short"],
                "publication_journal": NATURE_2025["journal"],
                "publication_year": NATURE_2025["year"],
                "publication_published_date": NATURE_2025["published_date"],
                "publication_issue_pages": NATURE_2025["issue"],
                "internal_status": "needs_value_crosscheck",
            }
        )
    return record


def finalize_units(record: dict[str, Any]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []

    density = to_float(record.get("density_g_cm3_raw"))
    if density is not None:
        record["density_kg_m3"] = density * 1000
        record["specific_volume_cm3_g"] = 1 / density if density else None
        record["specific_volume_m3_kg"] = (1 / density) * 1e-3 if density else None

    diameter = to_float(record.get("diameter_um_raw"))
    if diameter is not None:
        record["diameter_m"] = diameter * 1e-6

    linear_density = to_float(record.get("linear_density_tex_raw"))
    if linear_density is not None:
        record["linear_density_kg_m"] = linear_density * 1e-6

    tenacity = to_float(record.get("tenacity_N_tex_raw"))
    if tenacity is not None:
        record["specific_strength_N_m_kg"] = tenacity * 1e6

    tensile_raw = to_float(record.get("tensile_strength_GPa_raw"))
    if tensile_raw is not None:
        if tensile_raw > 100:
            tensile_gpa = tensile_raw / 1000
            record["tensile_strength_unit_inference"] = "raw_value_assumed_MPa_due_to_scale"
            issues.append(issue(record, "unit_inference", "tensile_strength_GPa_raw", f"Raw tensile value {tensile_raw} is >100 under GPa label; stored canonical value as MPa/1000."))
        else:
            tensile_gpa = tensile_raw
            record["tensile_strength_unit_inference"] = "raw_value_as_GPa"
        record["tensile_strength_GPa"] = tensile_gpa
        record["tensile_strength_Pa"] = tensile_gpa * 1e9

    modulus_n_tex = to_float(record.get("initial_modulus_N_tex_raw"))
    if modulus_n_tex is not None:
        record["specific_modulus_N_m_kg"] = modulus_n_tex * 1e6

    modulus_gpa = to_float(record.get("initial_modulus_GPa_raw"))
    if modulus_gpa is not None:
        record["initial_modulus_GPa"] = modulus_gpa
        record["initial_modulus_Pa"] = modulus_gpa * 1e9

    strain_pct = to_float(record.get("breaking_strain_pct_raw"))
    if strain_pct is not None:
        record["breaking_strain_fraction"] = strain_pct / 100

    wor = to_float(record.get("rupture_work_J_g_raw"))
    if wor is not None:
        record["rupture_work_J_kg"] = wor * 1000

    econd_s_cm = to_float(record.get("electrical_conductivity_S_cm_raw"))
    if econd_s_cm is not None:
        record["electrical_conductivity_S_m"] = econd_s_cm * 100
        if density:
            record["specific_electrical_conductivity_S_m2_kg"] = (econd_s_cm * 100) / (density * 1000)

    econd_ms_m = to_float(record.get("electrical_conductivity_MS_m_raw"))
    if econd_ms_m is not None:
        record["electrical_conductivity_S_m"] = econd_ms_m * 1e6
        if density:
            record["specific_electrical_conductivity_S_m2_kg"] = (econd_ms_m * 1e6) / (density * 1000)

    econd_specific_ms_m2_g = to_float(record.get("specific_electrical_conductivity_MS_m2_g_raw"))
    if econd_specific_ms_m2_g is not None:
        # The radar field is numerically electrical conductivity in MS/m divided
        # by density in g/cm3. Convert that ratio to sigma/rho in S m2/kg.
        record["specific_electrical_conductivity_S_m2_kg"] = econd_specific_ms_m2_g * 1000

    tcond = to_float(record.get("thermal_conductivity_W_mK_raw"))
    if tcond is not None:
        record["thermal_conductivity_W_mK"] = tcond
        if density:
            record["specific_thermal_conductivity_W_m2_K_kg"] = tcond / (density * 1000)

    gd = to_float(record.get("g_d_ratio_raw"))
    if gd is not None:
        record["g_d_ratio"] = gd

    ampacity = to_float(record.get("ampacity_A_cm2_raw"))
    if ampacity is not None:
        record["ampacity_A_m2"] = ampacity * 1e4

    ampacity_m2 = to_float(record.get("ampacity_A_m2_raw"))
    if ampacity_m2 is not None:
        record["ampacity_A_m2"] = ampacity_m2

    for doi in extract_dois(record.get("citation_raw")):
        if doi.startswith("10.1126/science.122806") and doi != "10.1126/science.1228061":
            issues.append(issue(record, "doi_validation", "doi_raw", f"Suspicious sequential DOI-like string `{doi}`; requires validation."))

    if not record.get("doi_raw") and record["material_family"] in {"CNT_or_CNT_hybrid", "graphene_or_GO_fiber", "CNT_metal_composite"}:
        issues.append(issue(record, "missing_reference", "doi_raw", "CNT/graphene row lacks DOI-like citation string."))

    if record["comparison_scope"] == "strict_candidate_missing_conditions":
        condition_fields = ["condition_temperature_C", "condition_atmosphere", "measurement_method", "gauge_length_mm"]
        missing = [field for field in condition_fields if not has_reported_value(record.get(field))]
        has_mechanical = any(record.get(field) is not None for field in ["tenacity_N_tex_raw", "tensile_strength_GPa_raw", "initial_modulus_GPa_raw", "initial_modulus_N_tex_raw"])
        if has_mechanical and not has_reported_value(record.get("strain_rate_s_inv")):
            missing.append("strain_rate_s_inv")
        if missing:
            issues.append(issue(record, "missing_conditions", "conditions", "Strict comparison candidate lacks structured metadata: " + ", ".join(missing) + "."))

    return issues


def has_reported_value(value: Any) -> bool:
    text = clean(value)
    if text is None:
        return False
    return str(text).strip().lower() not in {"not specified", "not reported", "unknown", "na", "n/a"}


def issue(record: dict[str, Any], issue_type: str, field: str, message: str) -> dict[str, Any]:
    return {
        "issue_id": stable_id("issue", record["record_id"], issue_type, field, message),
        "record_id": record["record_id"],
        "source_file": record["source_file"],
        "source_sheet": record["source_sheet"],
        "source_row": record["source_row"],
        "issue_type": issue_type,
        "field": field,
        "message": message,
    }


def read_xiao() -> list[dict[str, Any]]:
    wb = load_workbook(ROOT / "XiaO_DATA.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    records = []
    group = None
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=4, max_col=25, values_only=True), start=4):
        values = [clean(value) for value in row_values]
        if values[0] is not None:
            group = values[0]
        if not any(to_float(v) is not None for v in values[3:]):
            continue
        raw = dict(zip(XIAO_COLUMNS, values))
        record = base_record("XiaO_DATA.xlsx", "Sheet1", row_idx, raw["record_label"], raw["notes"], raw["citation_raw"], group)
        record.update(raw)
        record["source_citation_raw"] = raw["citation_raw"]
        citation_text = correct_known_citation_errors(norm_text(raw["citation_raw"]), record["material_family"])
        dois = sorted(set(extract_dois(citation_text) + doi_overrides_for(citation_text)))
        urls = extract_urls(citation_text)
        source_citation_class, evidence_tier, comparison_note = evidence_classification(record["material_family"], citation_text, dois, urls)
        record.update(
            {
                "citation_raw": citation_text or None,
                "doi_raw": ";".join(dois) if dois else None,
                "url_raw": ";".join(urls) if urls else None,
                "source_citation_class": source_citation_class,
                "evidence_tier": evidence_tier,
                "comparison_note": comparison_note,
            }
        )
        record["raw_payload_json"] = json.dumps(raw, ensure_ascii=False)
        records.append(record)
    wb.close()
    return records


def read_radar_sheet(sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(ROOT / "RadarFigureSource.xlsx", read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = dedupe_headers(list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))
    records = []
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean(value) for value in row_values]
        if not any(v is not None for v in values):
            continue
        raw = dict(zip(headers, values))
        reason = radar_exclusion_reason(raw)
        if reason:
            EXCLUDED_ROWS.append(excluded_row("RadarFigureSource.xlsx", sheet_name, row_idx, reason, raw))
            continue
        label = raw.get("Name")
        notes = raw.get("Comments")
        citation = raw.get("Publication") or raw.get("Unnamed: 14") or raw.get("Unnamed: 18") or raw.get("Unnamed: 19")
        record = base_record("RadarFigureSource.xlsx", sheet_name, row_idx, label, notes, citation)
        if sheet_name == "Radar":
            record.update(
                {
                    "density_g_cm3_raw": raw.get("Density [g cm−3] ±"),
                    "specific_volume_cm3_g_raw": raw.get("Specific Volume [cm3g-1]"),
                    "tenacity_N_tex_raw": raw.get("Tenacity  [N tex−1 = GPa/SG = GPa/(g/cm^3)] "),
                    "initial_modulus_N_tex_raw": raw.get("Initial Modulus [N tex−1 = GPa/SG = GPa/(g/cm^3)] "),
                    "rupture_work_J_g_raw": raw.get("Work of Rupture [J g−1]"),
                    "specific_electrical_conductivity_MS_m2_g_raw": raw.get("E Cond [MS m^2/g]"),
                    "thermal_conductivity_W_mK_raw": raw.get("Thermal Conductivity [W m−1 K−1]"),
                    "electrical_conductivity_MS_m_raw": raw.get("Electrical Conductivity [MS m−1]"),
                    "tensile_strength_GPa_raw": raw.get("Tensile [GPa]"),
                    "g_d_ratio_raw": raw.get("G:D Ratio"),
                }
            )
        else:
            record.update(
                {
                    "density_g_cm3_raw": raw.get("Density [g cm−3] ±"),
                    "specific_volume_cm3_g_raw": raw.get("1/Density [cm3g-1]"),
                    "density_g_cm3_error_raw": raw.get("Unnamed: 3"),
                    "tenacity_N_tex_raw": raw.get("Tenacity [N tex−1]"),
                    "tenacity_N_tex_error_raw": raw.get("±"),
                    "initial_modulus_N_tex_raw": raw.get("Initial Modulus [N tex−1]"),
                    "initial_modulus_N_tex_error_raw": raw.get("±.1"),
                    "breaking_strain_pct_raw": raw.get("Elongation [%]"),
                    "breaking_strain_pct_error_raw": raw.get("±.2"),
                    "rupture_work_J_g_raw": raw.get("Work of Rupture [J g−1]"),
                    "rupture_work_J_g_error_raw": raw.get("±.3"),
                    "electrical_conductivity_MS_m_raw": raw.get("Electrical Conductivity [MS m−1]"),
                    "electrical_conductivity_MS_m_error_raw": raw.get("±.4"),
                    "thermal_conductivity_W_mK_raw": raw.get("Thermal Conductivity [W m−1 K−1]"),
                    "thermal_conductivity_W_mK_error_raw": raw.get("±.5"),
                    "tensile_strength_GPa_raw": raw.get("Tensile (Gpa)"),
                }
            )
        record["raw_payload_json"] = json.dumps(raw, ensure_ascii=False, default=str)
        records.append(record)
    wb.close()
    return records


def read_annual_progression() -> pd.DataFrame:
    wb = load_workbook(ROOT / "RadarFigureSource.xlsx", read_only=True, data_only=True)
    ws = wb["FibersStrengthYr"]
    headers = dedupe_headers(list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))
    rows: list[dict[str, Any]] = []
    series_names = {
        "CNTF_SPSTR": "CNT fiber specific strength",
        "GF_SPSTR": "graphene fiber specific strength",
        "CNTSH_SPSTR": "CNT sheet specific strength",
        "GSH_SPSTR": "graphene sheet specific strength",
    }
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean(value) for value in row_values]
        raw = dict(zip(headers, values))
        year = to_float(raw.get("Year"))
        if year is None:
            continue
        for header, value in raw.items():
            if header == "Year" or str(header).startswith("Unnamed"):
                continue
            numeric = to_float(value)
            if numeric is None:
                continue
            base_header = str(header).split(".")[0]
            rows.append(
                {
                    "series_id": stable_id("series", "RadarFigureSource.xlsx", "FibersStrengthYr", row_idx, header),
                    "source_file": "RadarFigureSource.xlsx",
                    "source_sheet": "FibersStrengthYr",
                    "source_row": row_idx,
                    "year": int(year),
                    "series_code": header,
                    "series_label": series_names.get(base_header, base_header),
                    "property": "specific_strength",
                    "value_raw": numeric,
                    "unit_raw": "GPa/SG_or_N/tex_from_plot",
                    "value_canonical": numeric * 1e6,
                    "unit_canonical": "N m/kg",
                    "extraction_method": "derived_plot_workbook",
                    "raw_payload_json": json.dumps(raw, ensure_ascii=False, default=str),
                }
            )
    wb.close()
    return pd.DataFrame(rows)


def read_new_g() -> list[dict[str, Any]]:
    wb = load_workbook(ROOT / "New G fibre table Juan.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    records = []
    headers = [clean(value) for value in next(ws.iter_rows(min_row=3, max_row=3, min_col=2, max_col=7, values_only=True))]
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=4, min_col=2, max_col=7, values_only=True), start=4):
        values = [clean(value) for value in row_values]
        if not any(v is not None for v in values):
            continue
        raw = dict(zip(headers, values))
        citation = raw.get("Ref")
        notes = raw.get("Notes")
        label = "Graphene Fibers"
        if citation and "to be published" in lower_text(citation):
            notes = norm_text(notes, "domain-folded graphene fibre; Nature Materials publication verified")
        record = base_record("New G fibre table Juan.xlsx", "Sheet1", row_idx, label, notes, citation, "Graphene Fibers")
        record.update(
            {
                "density_g_cm3_raw": raw.get("Density, g/cc"),
                "tenacity_N_tex_raw": raw.get("Specific Strength, Gpa/SG"),
                "tensile_strength_GPa_raw": raw.get("Strength, GPa"),
                "publication_year_raw": raw.get("Year"),
            }
        )
        record["raw_payload_json"] = json.dumps(raw, ensure_ascii=False, default=str)
        records.append(record)
    wb.close()
    return records


def read_literature_addendum() -> list[dict[str, Any]]:
    path = ROOT / "data" / "literature" / "literature_addendum_records.tsv"
    if not path.exists():
        return []
    df = pd.read_csv(path, sep="\t", dtype=object).where(pd.notna, None)
    records = []
    for source_row, raw in enumerate(df.to_dict(orient="records"), start=2):
        label = raw.get("record_label")
        sample = raw.get("sample_name")
        citation = raw.get("citation_raw")
        group = raw.get("group_label")
        source_sheet = clean(raw.get("source_sheet")) or "manual_extract"
        record = base_record("literature_addendum_records.tsv", source_sheet, source_row, label, sample, citation, group)
        explicit_material = clean(raw.get("material_family")) or record["material_family"]
        explicit_form = clean(raw.get("form_factor")) or record["form_factor"]
        citation_text = correct_known_citation_errors(norm_text(citation), explicit_material)
        dois = sorted(set(extract_dois(citation_text) + doi_overrides_for(citation_text)))
        urls = extract_urls(citation_text)
        source_citation_class, evidence_tier, comparison_note = evidence_classification(explicit_material, citation_text, dois, urls)
        record.update(
            {
                "record_label": clean(label),
                "group_label": clean(group),
                "sample_name": clean(sample) or clean(label),
                "citation_raw": citation_text or None,
                "source_citation_raw": clean(raw.get("source_citation_raw")) or citation_text or None,
                "doi_raw": ";".join(dois) if dois else None,
                "url_raw": ";".join(urls) if urls else None,
                "material_family": explicit_material,
                "form_factor": explicit_form,
                "cnt_type": clean(raw.get("cnt_type")),
                "synthesis_method": clean(raw.get("synthesis_method")),
                "postprocessing": clean(raw.get("postprocessing")),
                "condition_temperature_C": to_float(raw.get("condition_temperature_C")),
                "condition_atmosphere": clean(raw.get("condition_atmosphere")),
                "measurement_method": clean(raw.get("measurement_method")),
                "gauge_length_mm": to_float(raw.get("gauge_length_mm")),
                "strain_rate_s_inv": to_float(raw.get("strain_rate_s_inv")),
                "provenance_table_figure_page": clean(raw.get("provenance_table_figure_page")),
                "source_export": clean(raw.get("source_export")) or "manual literature addendum from local PDFs",
                "extraction_method": clean(raw.get("extraction_method")) or "manual_literature_extract",
                "comparison_scope": comparison_scope(explicit_material, explicit_form),
                "source_citation_class": source_citation_class,
                "evidence_tier": evidence_tier,
                "comparison_note": comparison_note,
                "public_status": clean(raw.get("public_status")) or "internal_seed_not_public",
                "internal_status": clean(raw.get("internal_status")) or "ready_for_manual_review",
                "curator": clean(raw.get("curator")) or "source_compilation",
                "confidence_score": to_float(raw.get("confidence_score")),
                "notes": clean(raw.get("notes")),
                "raw_payload_json": json.dumps(raw, ensure_ascii=False, default=str),
            }
        )
        for field in [
            "density_g_cm3_raw",
            "diameter_um_raw",
            "linear_density_tex_raw",
            "tenacity_N_tex_raw",
            "tensile_strength_GPa_raw",
            "initial_modulus_N_tex_raw",
            "initial_modulus_GPa_raw",
            "breaking_strain_pct_raw",
            "rupture_work_J_g_raw",
            "electrical_conductivity_S_cm_raw",
            "electrical_conductivity_MS_m_raw",
            "specific_electrical_conductivity_MS_m2_g_raw",
            "thermal_conductivity_W_mK_raw",
            "g_d_ratio_raw",
            "ampacity_A_cm2_raw",
            "ampacity_gauge_length_mm",
        ]:
            record[field] = to_float(raw.get(field))
        records.append(record)
    return records


def load_james_doi_lookup() -> dict[str, dict[str, Any]]:
    path = ROOT / "data" / "literature" / "james_reference_doi_lookup.csv"
    if not path.exists():
        return {}
    df = pd.read_csv(path, dtype=object).where(pd.notna, None)
    return {str(row["reference_raw"]).strip(): row for row in df.to_dict(orient="records") if clean(row.get("reference_raw"))}


def james_has_measurement(raw: dict[str, Any]) -> bool:
    fields = [
        "Conductivity (MSm-1)",
        "Density (g cm-3)",
        "Specific Conductivity (kS m2/kg)",
        "Raman G:D",
        "Tensile Strength (MPa)",
        "Specific Strength (N/Tex)",
        "Young's Modulus (GPa)",
        "Bulk Fiber Diameter (microns)",
        "CNT Diameter (nm)",
        "Plottable CNT Diameter (nm)",
        "Thermal Conductivity (W/(m K))",
        "Absolute J max (A/m2)",
        "Absolute J max (A/m2).1",
    ]
    return any(to_float(raw.get(field)) is not None for field in fields)


def james_record_from_row(raw: dict[str, Any], source_row: int, lookup_row: dict[str, Any], suffix: str = "") -> dict[str, Any]:
    reference = clean(raw.get("Reference"))
    category = clean(raw.get("Category"))
    notes = clean(raw.get("Notes"))
    year = clean(raw.get("Year")) or clean(lookup_row.get("resolved_year"))
    material_family = james_material_family(category, raw.get("Category 2"), reference, notes)
    form_factor = james_form_factor(category, reference, notes, material_family)
    sample = notes or norm_text(category, year) or reference
    label = norm_text(reference, suffix).strip()
    doi = clean(lookup_row.get("resolved_doi"))
    record = base_record(JAMES_META_2021["source_file"], "Main", source_row, label, sample, doi, category)
    record.update(
        {
            "record_id": stable_id("rec", JAMES_META_2021["source_file"], "Main", source_row, label, suffix),
            "record_label": label,
            "group_label": category,
            "sample_name": sample,
            "citation_raw": doi,
            "source_citation_raw": f"Original DOI resolved from James/Bulmer workbook reference: {doi}; compiled data source DOI: {JAMES_META_2021['doi']}",
            "doi_raw": doi,
            "url_raw": None,
            "secondary_source_doi_raw": JAMES_META_2021["doi"],
            "secondary_source_title": JAMES_META_2021["title"],
            "secondary_source_authors_short": JAMES_META_2021["authors_short"],
            "secondary_source_journal": JAMES_META_2021["journal"],
            "secondary_source_year": JAMES_META_2021["year"],
            "original_reference_raw": reference,
            "doi_resolution_status": clean(lookup_row.get("resolution_status")),
            "doi_resolution_score": to_float(lookup_row.get("match_score")),
            "publication_title": clean(lookup_row.get("resolved_title")),
            "publication_journal": clean(lookup_row.get("resolved_journal")),
            "publication_year": to_float(lookup_row.get("resolved_year")),
            "material_family": material_family,
            "form_factor": form_factor,
            "cnt_type": james_cnt_type(category, reference, notes),
            "synthesis_method": james_synthesis(raw, material_family),
            "postprocessing": james_postprocessing(raw),
            "provenance_table_figure_page": "Bulmer/James meta-analysis supporting workbook, Main sheet",
            "source_export": JAMES_META_2021["source_export"],
            "extraction_method": "peer_reviewed_meta_analysis_workbook",
            "comparison_scope": comparison_scope(material_family, form_factor),
            "source_citation_class": "peer_reviewed_meta_analysis_record",
            "evidence_tier": "secondary_meta_analysis_with_original_doi",
            "comparison_note": "Peer-reviewed meta-analysis workbook row with original DOI resolved by Crossref/OpenAlex; not re-extracted from the primary paper in this pipeline.",
            "public_status": "secondary_public_candidate",
            "internal_status": "needs_primary_value_crosscheck",
            "curator": "James_Bulmer_meta_analysis_workbook",
            "confidence_score": 3,
            "raw_payload_json": json.dumps(raw, ensure_ascii=False, default=str),
        }
    )
    record["density_g_cm3_raw"] = to_float(raw.get("Density (g cm-3)"))
    record["specific_electrical_conductivity_MS_m2_g_raw"] = to_float(raw.get("Specific Conductivity (kS m2/kg)"))
    record["electrical_conductivity_MS_m_raw"] = to_float(raw.get("Conductivity (MSm-1)"))
    record["tenacity_N_tex_raw"] = to_float(raw.get("Specific Strength (N/Tex)"))
    tensile_mpa = to_float(raw.get("Tensile Strength (MPa)"))
    record["tensile_strength_GPa_raw"] = tensile_mpa / 1000 if tensile_mpa is not None else None
    record["initial_modulus_GPa_raw"] = to_float(raw.get("Young's Modulus (GPa)"))
    record["thermal_conductivity_W_mK_raw"] = to_float(raw.get("Thermal Conductivity (W/(m K))"))
    record["g_d_ratio_raw"] = to_float(raw.get("Raman G:D"))
    if form_factor == "individual_nanotube_or_bundle":
        cnt_diameter_nm = to_float(raw.get("Plottable CNT Diameter (nm)")) or to_float(raw.get("CNT Diameter (nm)"))
        record["diameter_um_raw"] = cnt_diameter_nm / 1000 if cnt_diameter_nm is not None else None
    else:
        record["diameter_um_raw"] = to_float(raw.get("Bulk Fiber Diameter (microns)"))
    electrical_probe_um = to_float(raw.get("Probe separation for Electrical conductivity (microns)"))
    if electrical_probe_um is not None:
        record["gauge_length_mm"] = electrical_probe_um / 1000
    method_parts = []
    if clean(raw.get("Alignment method")):
        method_parts.append(f"alignment: {raw.get('Alignment method')}")
    if clean(raw.get("Ampacity Measurement Config")):
        method_parts.append(f"ampacity: {raw.get('Ampacity Measurement Config')}")
    if electrical_probe_um is not None:
        method_parts.append(f"electrical probe separation {electrical_probe_um:g} um")
    record["measurement_method"] = "; ".join(method_parts) if method_parts else None
    return record


def read_james_metaanalysis() -> list[dict[str, Any]]:
    path = ROOT / "data" / "literature" / "source_files" / JAMES_META_2021["source_file"]
    if not path.exists():
        return []
    lookup = load_james_doi_lookup()
    df = pd.read_excel(path, sheet_name="Main", dtype=object).where(pd.notna, None)
    records: list[dict[str, Any]] = []
    for source_row, raw in enumerate(df.to_dict(orient="records"), start=2):
        reference = clean(raw.get("Reference"))
        if not reference or not james_has_measurement(raw):
            continue
        lookup_row = lookup.get(reference)
        if not lookup_row or clean(lookup_row.get("resolution_status")) != "accepted" or not clean(lookup_row.get("resolved_doi")):
            reason = "james_reference_doi_unresolved"
            if lookup_row and clean(lookup_row.get("resolution_status")) == "web_or_specsheet_reference":
                reason = "james_web_or_specsheet_reference_not_public"
            EXCLUDED_ROWS.append(excluded_row(JAMES_META_2021["source_file"], "Main", source_row, reason, raw))
            continue
        record = james_record_from_row(raw, source_row, lookup_row)
        record["ampacity_A_m2_raw"] = to_float(raw.get("Absolute J max (A/m2)"))
        ampacity_probe_um = to_float(raw.get("Probe separation for Ampacity (microns)"))
        if ampacity_probe_um is not None:
            record["ampacity_gauge_length_mm"] = ampacity_probe_um / 1000
        if clean(raw.get("Medium 1")):
            record["condition_atmosphere"] = clean(raw.get("Medium 1"))
        records.append(record)

        ampacity_2 = to_float(raw.get("Absolute J max (A/m2).1"))
        if ampacity_2 is not None:
            second = james_record_from_row(raw, source_row, lookup_row, suffix="secondary ampacity condition")
            for field in list(second):
                if field.endswith("_raw") or field in {
                    "density_kg_m3",
                    "specific_volume_cm3_g",
                    "specific_volume_m3_kg",
                    "diameter_m",
                    "linear_density_kg_m",
                    "specific_strength_N_m_kg",
                    "tensile_strength_GPa",
                    "tensile_strength_Pa",
                    "specific_modulus_N_m_kg",
                    "initial_modulus_GPa",
                    "initial_modulus_Pa",
                    "breaking_strain_fraction",
                    "rupture_work_J_kg",
                    "electrical_conductivity_S_m",
                    "specific_electrical_conductivity_S_m2_kg",
                    "thermal_conductivity_W_mK",
                    "specific_thermal_conductivity_W_m2_K_kg",
                    "g_d_ratio",
                }:
                    second[field] = None
            second["ampacity_A_m2_raw"] = ampacity_2
            if ampacity_probe_um is not None:
                second["ampacity_gauge_length_mm"] = ampacity_probe_um / 1000
                second["gauge_length_mm"] = ampacity_probe_um / 1000
            second["condition_atmosphere"] = clean(raw.get("Medium 2"))
            second["measurement_method"] = clean(raw.get("Ampacity Measurement Config")) or "ampacity"
            second["sample_name"] = norm_text(second["sample_name"], clean(raw.get("Medium 2")) or "secondary ampacity condition")
            records.append(second)
    return records


def close_enough(a: Any, b: Any) -> bool:
    av = to_float(a)
    bv = to_float(b)
    if av is None or bv is None:
        return False
    scale = max(abs(av), abs(bv), 1.0)
    return abs(av - bv) <= 0.025 * scale


def detect_secondary_duplicates(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    primary_records = [record for record in records if record.get("source_file") != JAMES_META_2021["source_file"]]
    for record in records:
        if record.get("source_file") != JAMES_META_2021["source_file"] or not record.get("doi_raw"):
            continue
        record_dois = set(extract_dois(record.get("doi_raw")))
        best_match: tuple[dict[str, Any], int, int] | None = None
        for other in primary_records:
            other_dois = set(extract_dois(other.get("doi_raw")))
            if not record_dois.intersection(other_dois):
                continue
            overlaps = [field for field in CANONICAL_DUPLICATE_FIELDS if record.get(field) is not None and other.get(field) is not None]
            if len(overlaps) < 2:
                continue
            matches = sum(1 for field in overlaps if close_enough(record.get(field), other.get(field)))
            if matches >= 2 and matches / len(overlaps) >= 0.75:
                if best_match is None or matches > best_match[1]:
                    best_match = (other, matches, len(overlaps))
        if best_match:
            other, matches, overlaps = best_match
            score = matches / overlaps
            record["duplicate_of_record_id"] = other["record_id"]
            record["duplicate_match_score"] = round(score, 4)
            record["internal_status"] = "duplicate_candidate_lower_priority_secondary_source"
            issues.append(
                issue(
                    record,
                    "duplicate_candidate",
                    "duplicate_of_record_id",
                    f"James/Bulmer secondary row is a likely duplicate of higher-priority record {other['record_id']} ({matches}/{overlaps} matching canonical properties).",
                )
            )
    return issues


def build_measurements(records: list[dict[str, Any]]) -> pd.DataFrame:
    definitions = [
        ("density", "density_kg_m3", "kg/m^3"),
        ("specific_volume", "specific_volume_m3_kg", "m^3/kg"),
        ("diameter", "diameter_m", "m"),
        ("linear_density", "linear_density_kg_m", "kg/m"),
        ("specific_strength", "specific_strength_N_m_kg", "N m/kg"),
        ("tensile_strength", "tensile_strength_Pa", "Pa"),
        ("specific_modulus", "specific_modulus_N_m_kg", "N m/kg"),
        ("initial_modulus", "initial_modulus_Pa", "Pa"),
        ("breaking_strain", "breaking_strain_fraction", "1"),
        ("work_of_rupture", "rupture_work_J_kg", "J/kg"),
        ("electrical_conductivity", "electrical_conductivity_S_m", "S/m"),
        ("specific_electrical_conductivity", "specific_electrical_conductivity_S_m2_kg", "S m^2/kg"),
        ("thermal_conductivity", "thermal_conductivity_W_mK", "W/m/K"),
        ("specific_thermal_conductivity", "specific_thermal_conductivity_W_m2_K_kg", "W m^2/K/kg"),
        ("g_d_ratio", "g_d_ratio", "ratio"),
        ("ampacity", "ampacity_A_m2", "A/m^2"),
    ]
    rows = []
    for record in records:
        for prop, field, unit in definitions:
            value = record.get(field)
            if value is None or pd.isna(value):
                continue
            rows.append(
                {
                    "measurement_id": stable_id("meas", record["record_id"], prop, field),
                    "record_id": record["record_id"],
                    "property": prop,
                    "value_canonical": value,
                    "unit_canonical": unit,
                    "source_file": record["source_file"],
                    "source_sheet": record["source_sheet"],
                    "source_row": record["source_row"],
                    "material_family": record.get("material_family"),
                    "form_factor": record.get("form_factor"),
                    "comparison_scope": record["comparison_scope"],
                    "source_citation_class": record.get("source_citation_class"),
                    "evidence_tier": record.get("evidence_tier"),
                    "comparison_note": record.get("comparison_note"),
                }
            )
    return pd.DataFrame(rows)


def build_publications(records: list[dict[str, Any]]) -> pd.DataFrame:
    pubs: dict[str, dict[str, Any]] = {}
    for record in records:
        dois = extract_dois(record.get("citation_raw"))
        urls = extract_urls(record.get("citation_raw"))
        if not dois and not urls and not record.get("citation_raw"):
            continue
        if not dois and record.get("doi_raw"):
            dois = str(record["doi_raw"]).split(";")
        if not urls and record.get("url_raw"):
            urls = str(record["url_raw"]).split(";")
        secondary_doi = clean(record.get("secondary_source_doi_raw"))
        if secondary_doi and secondary_doi not in dois:
            dois.append(secondary_doi)
        keys = dois or urls or [record.get("citation_raw")]
        for key in keys:
            if not key:
                continue
            if key == NATURE_2025["url"]:
                key = NATURE_2025["doi"]
            pub_id = stable_id("pub", key)
            existing = pubs.setdefault(
                pub_id,
                {
                    "publication_id": pub_id,
                    "doi": key if key.startswith("10.") else None,
                    "url": key if key.startswith("http") or key.startswith("www.") else None,
                    "title": None,
                    "authors_short": None,
                    "journal": None,
                    "year": None,
                    "published_date": None,
                    "issue_pages": None,
                    "citation_raw_examples": [],
                    "validation_status": "unvalidated",
                    "source_record_count": 0,
                },
            )
            if key == NATURE_2025["doi"] or key == NATURE_2025["url"]:
                existing.update(
                    {
                        "doi": NATURE_2025["doi"],
                        "url": NATURE_2025["url"],
                        "title": NATURE_2025["title"],
                        "authors_short": NATURE_2025["authors_short"],
                        "journal": NATURE_2025["journal"],
                        "year": NATURE_2025["year"],
                        "published_date": NATURE_2025["published_date"],
                        "issue_pages": NATURE_2025["issue"],
                        "validation_status": "verified_from_nature_page",
                    }
                )
            existing["source_record_count"] += 1
            raw = record.get("citation_raw")
            if raw and raw not in existing["citation_raw_examples"] and len(existing["citation_raw_examples"]) < 3:
                existing["citation_raw_examples"].append(raw)
    for row in pubs.values():
        row["citation_raw_examples"] = " || ".join(row["citation_raw_examples"])
    return pd.DataFrame(pubs.values()).sort_values(["validation_status", "doi", "url"], na_position="last")


def data_dictionary() -> pd.DataFrame:
    rows = [
        ("record_id", "Stable row identifier for a source record.", ""),
        ("source_file/source_sheet/source_row", "Exact provenance of extracted row.", ""),
        ("citation_raw", "Normalized/corrected citation text used for DOI extraction and publication grouping.", ""),
        ("source_citation_raw", "Original citation text from the source workbook when preserved separately from normalized citation text.", ""),
        ("secondary_source_*", "Secondary compilation source metadata, used when values entered through a peer-reviewed meta-analysis workbook rather than direct primary-paper extraction.", ""),
        ("original_reference_raw", "Original reference text from the secondary workbook before DOI resolution.", ""),
        ("doi_resolution_status/score", "Crossref/OpenAlex title-year resolution status and match score for secondary workbook references.", ""),
        ("duplicate_of_record_id", "Higher-priority record that appears to duplicate this lower-priority secondary row.", ""),
        ("*_raw", "Value as represented in source workbook after basic null cleaning.", "source unit"),
        ("density_kg_m3", "Canonical density.", "kg/m^3"),
        ("specific_volume_m3_kg", "Canonical specific volume calculated from density.", "m^3/kg"),
        ("tenacity_N_tex_raw", "Raw tenacity/specific strength from source.", "N/tex"),
        ("specific_strength_N_m_kg", "Canonical specific strength; 1 N/tex = 1e6 N m/kg.", "N m/kg"),
        ("tensile_strength_GPa", "Canonical tensile strength in GPa after scale inference.", "GPa"),
        ("tensile_strength_Pa", "Canonical tensile strength.", "Pa"),
        ("electrical_conductivity_S_m", "Canonical electrical conductivity; S/cm multiplied by 100 or MS/m by 1e6.", "S/m"),
        ("specific_electrical_conductivity_S_m2_kg", "Electrical conductivity divided by density.", "S m^2/kg"),
        ("thermal_conductivity_W_mK", "Thermal conductivity.", "W/m/K"),
        ("specific_thermal_conductivity_W_m2_K_kg", "Thermal conductivity divided by density.", "W m^2/K/kg"),
        ("ampacity_A_m2", "Maximum current density before open circuit/failure.", "A/m^2"),
        ("comparison_scope", "Initial comparison classification for UI filtering.", ""),
        ("source_citation_class", "Citation/provenance category used for public badges and source filters.", ""),
        ("evidence_tier", "Evidence strength tier distinguishing primary literature records from contextual benchmark comparators.", ""),
        ("comparison_note", "Human-readable disclosure text for website tooltips and figure legends.", ""),
        ("public_status", "Public exposure state. Seed rows are not marked public-official until curation/validation is complete.", ""),
        ("internal_status", "Internal curation state for reference, unit, condition, and value review.", ""),
        ("condition_*", "Nullable measurement-condition metadata. Missing values block strict scientific comparison.", ""),
        ("publication_year", "Formal citation year when known.", "year"),
        ("publication_published_date", "Online/version-of-record publication date when known.", "date"),
        ("annual_progression", "Separate derived plot-series table from RadarFigureSource/FibersStrengthYr.", ""),
        ("excluded_source_rows", "Rows deliberately excluded from scientific records because they are plot headers, max values, scores, or unlabeled plot cells.", ""),
    ]
    return pd.DataFrame(rows, columns=["field", "description", "canonical_unit"])


def main() -> None:
    EXCLUDED_ROWS.clear()
    OUT.mkdir(parents=True, exist_ok=True)
    records = []
    records.extend(read_literature_addendum())
    records.extend(read_xiao())
    records.extend(read_radar_sheet("Radar"))
    records.extend(read_radar_sheet("Fibers With Error"))
    records.extend(read_new_g())
    records.extend(read_james_metaanalysis())
    annual_progression_df = read_annual_progression()

    all_issues: list[dict[str, Any]] = []
    for record in records:
        all_issues.extend(finalize_units(record))
    all_issues.extend(detect_secondary_duplicates(records))

    records_df = pd.DataFrame(records)
    issues_df = pd.DataFrame(all_issues)
    measurements_df = build_measurements(records)
    publications_df = build_publications(records)
    dictionary_df = data_dictionary()
    excluded_df = pd.DataFrame(EXCLUDED_ROWS)
    duplicate_df = records_df[records_df.get("duplicate_of_record_id").notna()].copy() if "duplicate_of_record_id" in records_df else pd.DataFrame()

    records_df.to_csv(OUT / "combined_records.csv", index=False)
    measurements_df.to_csv(OUT / "measurements_long.csv", index=False)
    publications_df.to_csv(OUT / "publications.csv", index=False)
    issues_df.to_csv(OUT / "source_issues.csv", index=False)
    dictionary_df.to_csv(OUT / "data_dictionary.csv", index=False)
    annual_progression_df.to_csv(OUT / "annual_progression.csv", index=False)
    excluded_df.to_csv(OUT / "excluded_source_rows.csv", index=False)
    duplicate_df.to_csv(OUT / "duplicate_candidates.csv", index=False)

    sqlite_path = OUT / "cnt_property_database.sqlite"
    if sqlite_path.exists():
        sqlite_path.unlink()
    with sqlite3.connect(sqlite_path) as con:
        records_df.to_sql("records", con, index=False)
        measurements_df.to_sql("measurements", con, index=False)
        publications_df.to_sql("publications", con, index=False)
        issues_df.to_sql("source_issues", con, index=False)
        dictionary_df.to_sql("data_dictionary", con, index=False)
        annual_progression_df.to_sql("annual_progression", con, index=False)
        excluded_df.to_sql("excluded_source_rows", con, index=False)

    summary = {
        "records": len(records_df),
        "measurements": len(measurements_df),
        "publications": len(publications_df),
        "issues": len(issues_df),
        "annual_progression_points": len(annual_progression_df),
        "excluded_source_rows": len(excluded_df),
        "by_source": records_df["source_file"].value_counts().to_dict(),
        "by_material_family": records_df["material_family"].value_counts().to_dict(),
        "outputs": sorted(p.name for p in OUT.iterdir() if p.is_file()),
    }
    (OUT / "build_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
