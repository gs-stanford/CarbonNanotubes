#!/usr/bin/env python3
"""Inventory CNT review data assets into reproducible summaries."""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.io import loadmat


ROOT = Path(__file__).resolve().parents[1]
DOI_RE = re.compile(r"(10\.\d{4,9}/[-._;()/:A-Z0-9]+)", re.I)
URL_RE = re.compile(r"https?://\S+", re.I)
ASSIGN_RE = re.compile(
    r"(?P<var>[A-Za-z]\w*)\s*\(\s*(?P<row>\d+)\s*,\s*(?P<col>\d+)\s*\)\s*=\s*(?P<value>[-+0-9.eE]+)"
)
ARRAY_RE = re.compile(r"(?P<name>[A-Za-z]\w*)\s*=\s*\[(?P<values>[^\]]*)\]\s*;")
SCHEMA_RE = re.compile(r"%\s*(?P<name>[A-Za-z]\w*)\s+is\s+\[(?P<body>.*)", re.I)


PROPERTY_ALIASES = {
    "density": ("density", "rho"),
    "specific_volume": ("specific volume", "1/density"),
    "tenacity": ("tenacity", "specific strength", "spstr"),
    "initial_modulus": ("initial modulus", "specific modulus", "modulus", "stiffness"),
    "work_of_rupture": ("work of rupture", "rupture work", "energy to failure"),
    "electrical_conductivity": ("electrical conductivity", "e cond", "elcond"),
    "thermal_conductivity": ("thermal conductivity",),
    "tensile_strength": ("tensile", "strength"),
    "g_d_ratio": ("g:d", "g/d", "raman"),
    "diameter_thickness": ("diameter", "thickness"),
    "year": ("year",),
    "citation_or_doi": ("citation", "doi", "ref"),
    "comments": ("comments", "notes"),
}


MATLAB_VECTOR_SCHEMA = {
    "ArrayYarnVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "ArraySheetVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "ArrayLinkedVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "BuckyVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "FoamVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "ForestVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "SheetVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "YarnVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
    "WetSpunVec": [
        "density_kg_m3",
        "strength_mpa",
        "stiffness_gpa",
        "ductility_pct",
        "energy_to_failure_mj_m3",
        "electrical_conductivity_s_cm",
        "thermal_conductivity_w_mk",
    ],
}


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value in {"", "-", "--", "nan", "NaN"}:
            return None
    return value


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def property_matches(headers: list[str]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    lowered = {h: h.lower() for h in headers if h}
    for prop, aliases in PROPERTY_ALIASES.items():
        hits = [h for h, low in lowered.items() if any(alias in low for alias in aliases)]
        if hits:
            result[prop] = hits
    return result


def first_nonempty_header_row(path: Path, sheet_name: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name]
    best_row = 1
    best_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True):
        count = sum(clean_cell(v) is not None for v in row)
        if count > best_count:
            best_count = count
            best_row = row[0].row if hasattr(row[0], "row") else best_row
    # iter_rows(values_only=True) does not retain row indices.
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), start=1):
        count = sum(clean_cell(v) is not None for v in row)
        if count == best_count:
            wb.close()
            return idx
    wb.close()
    return best_row


def excel_summary(path: Path) -> dict[str, Any]:
    xls = pd.ExcelFile(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    out: dict[str, Any] = {"file": path.name, "sheets": []}
    for sheet in xls.sheet_names:
        ws = wb[sheet]
        header_row = first_nonempty_header_row(path, sheet)
        raw = pd.read_excel(path, sheet_name=sheet, header=header_row - 1, dtype=object)
        raw = raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
        headers = [norm_header(c) for c in raw.columns]
        raw.columns = headers
        missing_headers = sum(1 for h in headers if not h or h.startswith("Unnamed:"))
        rows = len(raw)
        col_summaries = []
        for col in headers:
            series = raw[col].map(clean_cell)
            non_null = int(series.notna().sum())
            doi_cells = int(series.astype(str).str.contains(DOI_RE, regex=True, na=False).sum())
            url_cells = int(series.astype(str).str.contains(URL_RE, regex=True, na=False).sum())
            numeric = pd.to_numeric(series, errors="coerce")
            num_count = int(numeric.notna().sum())
            col_summaries.append(
                {
                    "name": col,
                    "non_null": non_null,
                    "numeric": num_count,
                    "doi_cells": doi_cells,
                    "url_cells": url_cells,
                    "min": float(numeric.min()) if num_count else None,
                    "max": float(numeric.max()) if num_count else None,
                }
            )
        all_text = "\n".join(raw.astype(str).fillna("").to_numpy().ravel().tolist())
        dois = sorted(set(m.group(1).rstrip(" .,)") for m in DOI_RE.finditer(all_text)))
        out["sheets"].append(
            {
                "name": sheet,
                "dimension": ws.calculate_dimension(),
                "header_row": header_row,
                "rows": rows,
                "columns": headers,
                "missing_header_count": missing_headers,
                "property_matches": property_matches(headers),
                "doi_count_unique": len(dois),
                "doi_examples": dois[:10],
                "column_summaries": col_summaries,
            }
        )
    wb.close()
    return out


def extract_comment_references(lines: list[str]) -> list[str]:
    refs: list[str] = []
    buffer: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("%"):
            if buffer:
                text = " ".join(buffer)
                if '"' in text or DOI_RE.search(text) or re.search(r"\b(19|20)\d{2}\b", text):
                    refs.append(re.sub(r"\s+", " ", text).strip(" %"))
                buffer = []
            continue
        content = stripped.lstrip("%").strip()
        if not content or set(content) <= {"-", "_", "%"}:
            if buffer:
                text = " ".join(buffer)
                if '"' in text or DOI_RE.search(text) or re.search(r"\b(19|20)\d{2}\b", text):
                    refs.append(re.sub(r"\s+", " ", text).strip(" %"))
                buffer = []
            continue
        buffer.append(content)
    if buffer:
        text = " ".join(buffer)
        if '"' in text or DOI_RE.search(text) or re.search(r"\b(19|20)\d{2}\b", text):
            refs.append(re.sub(r"\s+", " ", text).strip(" %"))
    # Keep only likely citation chunks, deduplicate while preserving order.
    seen = set()
    cleaned = []
    for ref in refs:
        ref = ref.strip()
        if len(ref) < 20:
            continue
        key = ref.lower()
        if key not in seen:
            seen.add(key)
            cleaned.append(ref)
    return cleaned


def matlab_script_summary(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    vectors: dict[str, dict[str, Any]] = defaultdict(lambda: {"assigned_cells": [], "array_assignments": []})
    for m in ASSIGN_RE.finditer(text):
        var = m.group("var")
        row = int(m.group("row"))
        col = int(m.group("col"))
        value = float(m.group("value"))
        vectors[var]["assigned_cells"].append({"row": row, "col": col, "value": value})
    for m in ARRAY_RE.finditer(text):
        name = m.group("name")
        values_text = m.group("values")
        if "(" in name or name.lower() in {"zeros", "scatter", "legend", "xlabel", "ylabel", "xlim", "ylim"}:
            continue
        values = [
            float(v)
            for v in re.findall(r"[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?", values_text)
        ]
        if len(values) >= 2:
            vectors[name]["array_assignments"].append({"count": len(values), "min": min(values), "max": max(values)})
    vector_out = []
    for var, data in sorted(vectors.items()):
        cells = data["assigned_cells"]
        rows = sorted(set(c["row"] for c in cells))
        cols = sorted(set(c["col"] for c in cells))
        by_col: dict[int, list[float]] = defaultdict(list)
        for c in cells:
            by_col[c["col"]].append(c["value"])
        col_coverage = []
        for col in cols:
            values = by_col[col]
            schema = MATLAB_VECTOR_SCHEMA.get(var, [])
            col_coverage.append(
                {
                    "col": col,
                    "name": schema[col - 1] if col - 1 < len(schema) else None,
                    "count": len(values),
                    "min": min(values) if values else None,
                    "max": max(values) if values else None,
                }
            )
        vector_out.append(
            {
                "name": var,
                "assigned_cell_count": len(cells),
                "nonzero_rows": len(rows),
                "row_min": min(rows) if rows else None,
                "row_max": max(rows) if rows else None,
                "cols": cols,
                "col_coverage": col_coverage,
                "array_assignments": data["array_assignments"],
            }
        )
    comment_text = "\n".join(line.lstrip("%").strip() for line in lines if line.strip().startswith("%"))
    dois = sorted(set(m.group(1).rstrip(" .,)") for m in DOI_RE.finditer(comment_text)))
    return {
        "file": path.name,
        "line_count": len(lines),
        "vectors": vector_out,
        "comment_reference_count": len(extract_comment_references(lines)),
        "comment_reference_examples": extract_comment_references(lines)[:8],
        "doi_count_unique": len(dois),
        "doi_examples": dois[:10],
        "url_count": len(URL_RE.findall(comment_text)),
    }


def origin_binary_summary(path: Path) -> dict[str, Any]:
    raw = path.read_bytes()
    strings = re.findall(rb"[\x20-\x7E]{4,}", raw)
    decoded = [s.decode("latin1", errors="replace") for s in strings]
    joined = "\n".join(decoded)
    fields = [
        "Citation",
        "Diameter",
        "Density",
        "Linear Density",
        "Tenacity",
        "Tensile Strength",
        "Initial Modulus",
        "Breaking strain",
        "Rupture Work",
        "Electrical Conductivity",
        "Thermal Conductivity",
    ]
    graph_names = sorted(set(re.findall(r"[A-Za-z][A-Za-z0-9 -]{2,40}", joined)))
    dois = sorted(set(m.group(1).rstrip(" .,)") for m in DOI_RE.finditer(joined)))
    urls = URL_RE.findall(joined)
    return {
        "file": path.name,
        "bytes": len(raw),
        "ascii_string_count": len(decoded),
        "field_hits": [f for f in fields if f.lower() in joined.lower()],
        "doi_count_unique": len(dois),
        "doi_examples": dois[:20],
        "url_count": len(urls),
        "url_examples": urls[:10],
        "book_or_sheet_mentions": sorted(set(re.findall(r"Book\d+|Sheet\d+|Graph\d+|Folder\d+", joined)))[:50],
        "title_like_examples": [s for s in decoded if any(w in s.lower() for w in ["ashby", "thermal", "tenacity", "citation"])][:20],
        "token_examples": graph_names[:50],
    }


def fig_summary(path: Path) -> dict[str, Any]:
    mat = loadmat(path, squeeze_me=True, struct_as_record=False)
    keys = sorted([k for k in mat.keys() if not k.startswith("__")])
    return {"file": path.name, "keys": keys, "meta_repr": repr(mat.get("meta_data", ""))[:500]}


def main() -> None:
    output: dict[str, Any] = {
        "root": str(ROOT),
        "excel": [],
        "matlab_scripts": [],
        "origin": [],
        "figures": [],
    }
    for path in sorted(ROOT.glob("*.xlsx")):
        output["excel"].append(excel_summary(path))
    for path in sorted(ROOT.glob("*.m")):
        output["matlab_scripts"].append(matlab_script_summary(path))
    for path in sorted(ROOT.glob("*.opju")):
        output["origin"].append(origin_binary_summary(path))
    for path in sorted(ROOT.glob("*.fig")):
        output["figures"].append(fig_summary(path))
    print(json.dumps(output, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
